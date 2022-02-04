from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
import openpyxl

# Create a new workbook
wb = Workbook()
ws = wb.active

# Create new worksheets
ws1 = wb.create_sheet('New Sheet')
ws2 = wb.create_sheet('First', 0)

# Alter title of current worksheet
ws.title = 'MySheet'

print(wb.sheetnames)

# Create a 2nd workbook and load an existing excel file.
wb2 = load_workbook('Regions.xlsx')
new_sheet = wb2.create_sheet('NewSheet')
active_sheet = wb2.active

# Access the value at cell A1
cell = active_sheet['A1']
print(cell.value)
active_sheet['A1'] = 'Region'

# Save modified excel sheet
wb2.save('modified_regions.xlsx')

# Learning how to navigate through data
# Create a range of cells to navigate
wb = load_workbook('regions.xlsx')
ws = wb.active

cell_range = ws['A1':'C1']
col_range = ws['A': 'C']
print(col_range)

# Set a row range
row_range = ws[1:5]
print(row_range)

# Iterate through columns and rows.
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    for cell in row:
        print(cell)

wb.save('regions.xlsx')

wb = Workbook()
ws = wb.active

# Fill worksheet with integers
for i in range(1, 20):
    ws.append(range(300))

# Merge and then unmerge A1 to B5
ws.merge_cells('A1:B5')
ws.unmerge_cells('A1:B5')
ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)

cell = ws['B2']
cell.font = Font(color=colors.BLUE, size=20, italic=True)
cell.value = 'Merged Cell'
cell.alignment = Alignment(horizontal='right', vertical='bottom')
cell.fill = GradientFill(stop=('000000', 'FFFFFF'))

wb.save('text.xlsx')

# Storing styles
highlight = NamedStyle(name='highlight')
highlight.font = Font(bold=True)
bd = Side(style='thick', color='000000')
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = PatternFill('solid', fgColor='FFFF00')

# Apply this style to every cell DIAGONALLY, starting at the H column.
count = 0
for col in ws.iter_cols(min_col=8, min_row=1, max_col=30, max_row=30):
    col[count].style = highlight
    count += 1

wb.save('highlight.xlsx')

wb = openpyxl.Workbook()
ws = wb.active

# Create a simple data set
data = [['Food', 'Price'],
        ['Cheeseburger', '2.00'],
        ['Chicken Tendies', '5.00'],
        ['Chicken Caesar Salad', '8.00'],
        ['Pepperoni Pizza', '2.50']]

# Append this data set to the worksheet using a for loop.
for rows in data:
    ws.append(rows)
