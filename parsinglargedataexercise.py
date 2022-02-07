import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

wb = load_workbook('template.xlsx')
ws = wb.active

# Add a DataFrame. NOTE: Data set is LARGE. Add encoding flag (so it knows what encoding to read from,
# AND add casting to the data type of the file.)

df = pd.read_csv('crime.csv', encoding='utf-8', dtype={'INCIDENT_NUMBER': str,
                                                       'OFFENSE_CODE': str,
                                                       'OFFENSE_CODE_GROUP': str,
                                                       'OFFENSE_DESCRIPTION': str,
                                                       'DISTRICT': str,
                                                       'REPORTING_AREA': str,
                                                       'SHOOTING': str,
                                                       'YEAR': str,
                                                       'MONTH': str,
                                                       'DAY_OF_WEEK': str,
                                                       'HOUR': str})

# Search for offense code Counterfeiting
df1 = df[df['OFFENSE_CODE_GROUP'] == 'Counterfeiting']

# Use numpy to replace all NaN values
df1 = df1.replace(np.nan, 'N/A', regex=True)

# Extract some stats. Take total # of crimes, total number of counterfeiting crimes, and create a %.
total_crimes = len(df.index)
counterfeit = len(df1.index)
perc_crimes = (counterfeit / total_crimes) * 100
perc_crimes = round(perc_crimes, 2)

# Paste data to corresponding columns
ws['O8'].value = total_crimes
ws['P8'].value = counterfeit
ws['Q8'].value = perc_crimes

# Counterfeiting crimes occuring each year by district. Create a count column to measure.
# Use the unstack function to make it look clean.
df1['Count'] = 1
df2 = df1.groupby(['DISTRICT', 'YEAR']).count()['Count'].unstack(level=0)

# Drop the N/A Column
df2.drop(columns='N/A', inplace=True)

# Convert to openpyxl format
rows = dataframe_to_rows(df2)

# Add data to template starting at A8 using nested for loop (enumerate)
for r_idx, row in enumerate(rows, 8):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save('crime_report.xlsx')
