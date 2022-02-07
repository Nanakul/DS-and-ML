import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

df_1 = pd.read_excel('shifts.xlsx', sheet_name='Sheet')
df_2 = pd.read_excel('shifts.xlsx', sheet_name='Sheet1')
df_3 = pd.read_excel('shift_3.xlsx')

# Concatenate to combine all of the data frames
df_all = pd.concat([df_1, df_2, df_3], sort=False)

print(df_all)

to_excel = df_all.to_excel('all_shifts.xlsx', index=None)

wb = load_workbook('all_shifts.xlsx')
ws = wb.active

# Create new column called 'Total' with bold font in column G
total_col = ws['G1']
total_col.font = Font(bold=True)
total_col.value = 'Total'

# Create two variables to store strings for columns E and F
e_col, f_col = ['E', 'F']

# Create for loop to iterate through all the columns
for row in range(2,300):
    result_cell = 'G{}'.format(row)
    e_value = ws[e_col + str(row)].value
    f_value = ws[f_col + str(row)].value
    ws[result_cell] = e_value * f_value

wb.save('totaled.xlsx')

# Load in the regions.xlsx
wb = load_workbook('regions.xlsx')
ws = wb.active
df = pd.read_excel('all_shifts.xlsx')

# Create new DF to reflect Sales Rep, Cost per unit, and units sold
df1 = df[['Sales Rep', 'Cost per', 'Units Sold']]

# Create new Total Column
df1['Total'] = df1['Cost per'] * df1['Units Sold']
print(df1)
